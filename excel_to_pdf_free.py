#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel to PDF 자동 변환 프로그램 (무료 버전) v3.0
엑셀 파일의 링크를 자동으로 열고 PDF로 저장

주요 기능:
- 시트별 폴더 자동 생성
- 구글 계정 다중 관리
- 진행상황 실시간 표시
- 에러 로그 기록
- 완료 후 통계 제공
"""

import sys
import os
import time
import threading
import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from tkinter import Tk, Label, Button, Entry, Listbox, Checkbutton, IntVar, StringVar, Frame, Scrollbar, messagebox, filedialog, ttk, Text, Toplevel
from tkinter import MULTIPLE, END, VERTICAL, RIGHT, LEFT, BOTH, Y, X, TOP, BOTTOM, DISABLED, NORMAL
import base64
import logging
from datetime import datetime
import json
import psutil

class ExcelToPDFApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel to PDF 변환 프로그램 v3.0 (무료)")
        self.root.geometry("700x850")
        self.root.resizable(False, False)
        
        # 변수 초기화
        self.excel_path = None
        self.save_folder = None
        self.wb = None
        self.driver = None
        self.is_running = False
        self.is_paused = False
        self.total_processed = 0
        self.total_success = 0
        self.total_failed = 0
        self.failed_items = []
        self.start_time = None
        
        # 설정 디렉토리
        self.config_dir = os.path.join(os.path.expanduser("~"), "Documents", "Excel_to_PDF_Config")
        if not os.path.exists(self.config_dir):
            os.makedirs(self.config_dir)
        
        # 로그 설정
        self.setup_logging()
        
        # 구글 계정 로드
        self.google_accounts = self.load_google_accounts()
        
        self.setup_ui()
        
    def setup_logging(self):
        """로그 시스템 초기화"""
        log_dir = os.path.join(os.path.expanduser("~"), "Documents", "Excel_to_PDF_Logs")
        if not os.path.exists(log_dir):
            os.makedirs(log_dir)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        log_file = os.path.join(log_dir, f"conversion_{timestamp}.log")
        
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler(log_file, encoding='utf-8'),
                logging.StreamHandler()
            ]
        )
        
        self.logger = logging.getLogger(__name__)
        self.logger.info("=" * 50)
        self.logger.info("Excel to PDF 변환 프로그램 시작 (무료 버전)")
        self.logger.info("=" * 50)
        
    def load_google_accounts(self):
        """저장된 구글 계정 목록 로드"""
        accounts_file = os.path.join(self.config_dir, "google_accounts.json")
        if os.path.exists(accounts_file):
            try:
                with open(accounts_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except:
                return []
        return []
    
    def save_google_accounts(self):
        """구글 계정 목록 저장"""
        accounts_file = os.path.join(self.config_dir, "google_accounts.json")
        with open(accounts_file, 'w', encoding='utf-8') as f:
            json.dump(self.google_accounts, f, ensure_ascii=False, indent=2)
        
    def setup_ui(self):
        """UI 구성"""
        # 타이틀
        title_frame = Frame(self.root, bg="#4CAF50", height=70)
        title_frame.pack(fill=X)
        title_label = Label(title_frame, text="Excel → PDF 변환기 v3.0", 
                           font=("맑은 고딕", 20, "bold"), bg="#4CAF50", fg="white")
        title_label.pack(pady=20)
        
        # 메인 프레임
        main_frame = Frame(self.root, padx=20, pady=20)
        main_frame.pack(fill=BOTH, expand=True)
        
        # 1. 구글 계정 선택
        Label(main_frame, text="1. 구글 계정 선택", font=("맑은 고딕", 11, "bold")).pack(anchor="w", pady=(0,5))
        
        account_frame = Frame(main_frame)
        account_frame.pack(fill=X, pady=(0,15))
        
        self.account_var = StringVar()
        self.account_dropdown = ttk.Combobox(account_frame, textvariable=self.account_var, 
                                            state="readonly", width=40, font=("맑은 고딕", 9))
        self.account_dropdown.pack(side=LEFT, padx=(0,10))
        self.update_account_dropdown()
        
        Button(account_frame, text="계정 추가", command=self.add_google_account,
               bg="#2196F3", fg="white", font=("맑은 고딕", 9, "bold"),
               padx=10, pady=5, relief="flat", cursor="hand2").pack(side=LEFT, padx=(0,5))
        
        Button(account_frame, text="계정 삭제", command=self.delete_google_account,
               bg="#f44336", fg="white", font=("맑은 고딕", 9, "bold"),
               padx=10, pady=5, relief="flat", cursor="hand2").pack(side=LEFT)
        
        # 2. 엑셀 파일 선택
        Label(main_frame, text="2. 엑셀 파일 선택", font=("맑은 고딕", 11, "bold")).pack(anchor="w", pady=(0,5))
        file_frame = Frame(main_frame)
        file_frame.pack(fill=X, pady=(0,15))
        
        self.file_label = Label(file_frame, text="파일을 선택하세요", bg="#f0f0f0", 
                               anchor="w", padx=10, pady=8, relief="solid", borderwidth=1)
        self.file_label.pack(side=LEFT, fill=X, expand=True, padx=(0,10))
        
        Button(file_frame, text="파일 선택", command=self.select_file, 
               bg="#FF9800", fg="white", font=("맑은 고딕", 9, "bold"),
               padx=15, pady=5, relief="flat", cursor="hand2").pack(side=RIGHT)
        
        # 3. 열 선택
        Label(main_frame, text="3. URL이 있는 열 선택", font=("맑은 고딕", 11, "bold")).pack(anchor="w", pady=(0,5))
        col_frame = Frame(main_frame)
        col_frame.pack(fill=X, pady=(0,15))
        
        Label(col_frame, text="열 문자 (예: A, B, C):", font=("맑은 고딕", 9)).pack(side=LEFT, padx=(0,10))
        self.col_entry = Entry(col_frame, width=5, font=("맑은 고딕", 11), justify="center")
        self.col_entry.pack(side=LEFT)
        self.col_entry.insert(0, "A")
        
        # 4. 시트 선택
        Label(main_frame, text="4. 처리할 시트 선택", font=("맑은 고딕", 11, "bold")).pack(anchor="w", pady=(0,5))
        sheet_frame = Frame(main_frame)
        sheet_frame.pack(fill=X, pady=(0,15))
        
        list_frame = Frame(sheet_frame)
        list_frame.pack(fill=BOTH, expand=True)
        
        scrollbar = Scrollbar(list_frame, orient=VERTICAL)
        self.sheet_listbox = Listbox(list_frame, selectmode=MULTIPLE, height=6,
                                     yscrollcommand=scrollbar.set, font=("맑은 고딕", 9))
        scrollbar.config(command=self.sheet_listbox.yview)
        scrollbar.pack(side=RIGHT, fill=Y)
        self.sheet_listbox.pack(side=LEFT, fill=BOTH, expand=True)
        
        # 5. 저장 폴더 선택
        Label(main_frame, text="5. PDF 저장 폴더 선택", font=("맑은 고딕", 11, "bold")).pack(anchor="w", pady=(0,5))
        folder_frame = Frame(main_frame)
        folder_frame.pack(fill=X, pady=(0,15))
        
        self.folder_label = Label(folder_frame, text="폴더를 선택하세요", bg="#f0f0f0",
                                 anchor="w", padx=10, pady=8, relief="solid", borderwidth=1)
        self.folder_label.pack(side=LEFT, fill=X, expand=True, padx=(0,10))
        
        Button(folder_frame, text="폴더 선택", command=self.select_folder,
               bg="#FF9800", fg="white", font=("맑은 고딕", 9, "bold"),
               padx=15, pady=5, relief="flat", cursor="hand2").pack(side=RIGHT)
        
        # 진행 상태
        Label(main_frame, text="진행 상태", font=("맑은 고딕", 11, "bold")).pack(anchor="w", pady=(10,5))
        
        self.progress_label = Label(main_frame, text="대기 중...", 
                                   font=("맑은 고딕", 10), fg="#666")
        self.progress_label.pack(anchor="w")
        
        self.progress_bar = ttk.Progressbar(main_frame, length=300, mode='determinate')
        self.progress_bar.pack(fill=X, pady=(5,10))
        
        # 통계 정보
        stats_frame = Frame(main_frame, bg="#f5f5f5", relief="solid", borderwidth=1)
        stats_frame.pack(fill=X, pady=(0,15), padx=5)
        
        stat_row = Frame(stats_frame, bg="#f5f5f5")
        stat_row.pack(fill=X, pady=10, padx=10)
        
        self.stat_total = Label(stat_row, text="처리: 0", font=("맑은 고딕", 9), bg="#f5f5f5")
        self.stat_total.pack(side=LEFT, padx=10)
        
        self.stat_success = Label(stat_row, text="성공: 0", font=("맑은 고딕", 9), 
                                 bg="#f5f5f5", fg="#4CAF50")
        self.stat_success.pack(side=LEFT, padx=10)
        
        self.stat_failed = Label(stat_row, text="실패: 0", font=("맑은 고딕", 9), 
                                bg="#f5f5f5", fg="#f44336")
        self.stat_failed.pack(side=LEFT, padx=10)
        
        self.stat_time = Label(stat_row, text="경과: 00:00", font=("맑은 고딕", 9), bg="#f5f5f5")
        self.stat_time.pack(side=LEFT, padx=10)
        
        # 실행 버튼
        button_frame = Frame(main_frame)
        button_frame.pack(pady=10)
        
        self.start_btn = Button(button_frame, text="변환 시작", command=self.start_conversion,
                               bg="#4CAF50", fg="white", font=("맑은 고딕", 12, "bold"),
                               width=12, height=2, relief="flat", cursor="hand2")
        self.start_btn.pack(side=LEFT, padx=5)
        
        self.pause_btn = Button(button_frame, text="일시정지", command=self.toggle_pause,
                               bg="#FF9800", fg="white", font=("맑은 고딕", 12, "bold"),
                               width=12, height=2, relief="flat", cursor="hand2", state=DISABLED)
        self.pause_btn.pack(side=LEFT, padx=5)
        
        self.stop_btn = Button(button_frame, text="중지", command=self.stop_conversion,
                              bg="#f44336", fg="white", font=("맑은 고딕", 12, "bold"),
                              width=12, height=2, relief="flat", cursor="hand2", state=DISABLED)
        self.stop_btn.pack(side=LEFT, padx=5)
        
    def update_account_dropdown(self):
        """계정 드롭다운 업데이트"""
        if self.google_accounts:
            self.account_dropdown['values'] = self.google_accounts
            if not self.account_var.get():
                self.account_var.set(self.google_accounts[0])
        else:
            self.account_dropdown['values'] = []
            self.account_var.set("")
    
    def add_google_account(self):
        """구글 계정 추가"""
        account = simpledialog.askstring("계정 추가", "구글 계정 이메일을 입력하세요:")
        if account:
            if account not in self.google_accounts:
                self.google_accounts.append(account)
                self.save_google_accounts()
                self.update_account_dropdown()
                self.account_var.set(account)
                messagebox.showinfo("성공", f"계정이 추가되었습니다: {account}")
            else:
                messagebox.showwarning("중복", "이미 등록된 계정입니다.")
    
    def delete_google_account(self):
        """구글 계정 삭제"""
        current = self.account_var.get()
        if current and current in self.google_accounts:
            if messagebox.askyesno("확인", f"'{current}' 계정을 삭제하시겠습니까?"):
                self.google_accounts.remove(current)
                self.save_google_accounts()
                self.update_account_dropdown()
                messagebox.showinfo("성공", "계정이 삭제되었습니다.")
        else:
            messagebox.showwarning("경고", "삭제할 계정을 선택하세요.")
    
    def select_file(self):
        """엑셀 파일 선택"""
        file_path = filedialog.askopenfilename(
            title="엑셀 파일 선택",
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        
        if file_path:
            self.excel_path = file_path
            self.file_label.config(text=os.path.basename(file_path))
            self.logger.info(f"엑셀 파일 선택: {file_path}")
            
            # 엑셀 파일 열기
            try:
                self.wb = openpyxl.load_workbook(file_path, data_only=True)
                self.sheet_listbox.delete(0, END)
                for sheet_name in self.wb.sheetnames:
                    self.sheet_listbox.insert(END, sheet_name)
                messagebox.showinfo("성공", "엑셀 파일을 불러왔습니다!")
                self.logger.info(f"시트 목록: {self.wb.sheetnames}")
            except Exception as e:
                messagebox.showerror("오류", f"파일을 열 수 없습니다: {str(e)}")
                self.logger.error(f"파일 열기 실패: {str(e)}")
    
    def select_folder(self):
        """저장 폴더 선택"""
        folder_path = filedialog.askdirectory(title="PDF 저장 폴더 선택")
        if folder_path:
            self.save_folder = folder_path
            self.folder_label.config(text=folder_path)
            self.logger.info(f"저장 폴더 선택: {folder_path}")
    
    def toggle_pause(self):
        """일시정지/재개"""
        self.is_paused = not self.is_paused
        if self.is_paused:
            self.pause_btn.config(text="재개")
            self.progress_label.config(text="일시정지됨")
            self.logger.info("변환 일시정지")
        else:
            self.pause_btn.config(text="일시정지")
            self.progress_label.config(text="변환 재개...")
            self.logger.info("변환 재개")
    
    def stop_conversion(self):
        """변환 중지"""
        if messagebox.askyesno("확인", "변환을 중지하시겠습니까?"):
            self.is_running = False
            self.logger.info("사용자가 변환 중지")
    
    def start_conversion(self):
        """변환 시작"""
        # 입력 검증
        if not self.account_var.get():
            messagebox.showwarning("경고", "구글 계정을 선택하세요.")
            return
        
        if not self.excel_path:
            messagebox.showwarning("경고", "엑셀 파일을 선택하세요.")
            return
        
        if not self.sheet_listbox.curselection():
            messagebox.showwarning("경고", "처리할 시트를 선택하세요.")
            return
        
        if not self.save_folder:
            messagebox.showwarning("경고", "저장 폴더를 선택하세요.")
            return
        
        col = self.col_entry.get().strip().upper()
        if not col or not col.isalpha():
            messagebox.showwarning("경고", "올바른 열 문자를 입력하세요 (예: A, B, C).")
            return
        
        # 통계 초기화
        self.total_processed = 0
        self.total_success = 0
        self.total_failed = 0
        self.failed_items = []
        self.is_running = True
        self.is_paused = False
        self.start_time = time.time()
        
        # 버튼 상태 변경
        self.start_btn.config(state=DISABLED)
        self.pause_btn.config(state=NORMAL)
        self.stop_btn.config(state=NORMAL)
        
        # 별도 스레드에서 실행
        thread = threading.Thread(target=self.run_conversion, daemon=True)
        thread.start()
        
        # 타이머 시작
        self.update_timer()
    
    def update_timer(self):
        """경과 시간 업데이트"""
        if self.is_running and self.start_time:
            elapsed = int(time.time() - self.start_time)
            minutes = elapsed // 60
            seconds = elapsed % 60
            self.stat_time.config(text=f"경과: {minutes:02d}:{seconds:02d}")
            self.root.after(1000, self.update_timer)
    
    def run_conversion(self):
        """실제 변환 작업"""
        try:
            account = self.account_var.get()
            col = self.col_entry.get().strip().upper()
            selected_sheets = [self.sheet_listbox.get(i) for i in self.sheet_listbox.curselection()]
            
            self.logger.info(f"변환 시작 - 계정: {account}, 열: {col}, 시트: {selected_sheets}")
            
            # 크롬 드라이버 설정
            self.setup_chrome_driver(account)
            
            total_links = 0
            for sheet_name in selected_sheets:
                sheet = self.wb[sheet_name]
                for row in range(2, sheet.max_row + 1):
                    cell_value = sheet[f"{col}{row}"].value
                    if cell_value and str(cell_value).startswith("http"):
                        total_links += 1
            
            self.progress_bar['maximum'] = total_links
            self.logger.info(f"총 {total_links}개의 링크 발견")
            
            # 각 시트별로 처리
            for sheet_name in selected_sheets:
                if not self.is_running:
                    break
                
                # 시트별 폴더 생성
                sheet_folder = os.path.join(self.save_folder, sheet_name)
                if not os.path.exists(sheet_folder):
                    os.makedirs(sheet_folder)
                    self.logger.info(f"폴더 생성: {sheet_folder}")
                
                sheet = self.wb[sheet_name]
                self.logger.info(f"시트 '{sheet_name}' 처리 시작")
                
                for row in range(2, sheet.max_row + 1):
                    # 일시정지 체크
                    while self.is_paused and self.is_running:
                        time.sleep(0.5)
                    
                    if not self.is_running:
                        break
                    
                    cell_value = sheet[f"{col}{row}"].value
                    
                    if cell_value and str(cell_value).startswith("http"):
                        url = str(cell_value)
                        self.total_processed += 1
                        
                        # 진행 상태 업데이트
                        self.progress_label.config(
                            text=f"[{sheet_name}] {self.total_processed}/{total_links} 처리 중..."
                        )
                        self.progress_bar['value'] = self.total_processed
                        
                        # PDF 변환
                        success = self.convert_to_pdf(url, sheet_folder, f"row_{row}")
                        
                        if success:
                            self.total_success += 1
                            self.logger.info(f"✓ 성공: {url} → row_{row}.pdf")
                        else:
                            self.total_failed += 1
                            self.failed_items.append({
                                'sheet': sheet_name,
                                'row': row,
                                'url': url
                            })
                            self.logger.error(f"✗ 실패: {url}")
                        
                        # 통계 업데이트
                        self.stat_total.config(text=f"처리: {self.total_processed}")
                        self.stat_success.config(text=f"성공: {self.total_success}")
                        self.stat_failed.config(text=f"실패: {self.total_failed}")
                        
                        time.sleep(2)  # 안정성을 위한 대기
            
            # 완료 처리
            self.finish_conversion()
            
        except Exception as e:
            self.logger.error(f"변환 중 오류: {str(e)}")
            messagebox.showerror("오류", f"변환 중 오류가 발생했습니다: {str(e)}")
        finally:
            self.cleanup()
    
    def setup_chrome_driver(self, account):
        """크롬 드라이버 설정"""
        chrome_options = Options()
        
        # 프로필 디렉토리 설정
        profile_dir = os.path.join(os.path.expanduser("~"), "Documents", 
                                   "Excel_to_PDF_Profiles", account.replace("@", "_at_"))
        if not os.path.exists(profile_dir):
            os.makedirs(profile_dir)
        
        chrome_options.add_argument(f"user-data-dir={profile_dir}")
        chrome_options.add_argument("--disable-gpu")
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        
        # PDF 인쇄 설정
        prefs = {
            "printing.print_preview_sticky_settings.appState": json.dumps({
                "recentDestinations": [{
                    "id": "Save as PDF",
                    "origin": "local",
                    "account": ""
                }],
                "selectedDestinationId": "Save as PDF",
                "version": 2
            }),
            "savefile.default_directory": self.save_folder
        }
        chrome_options.add_experimental_option("prefs", prefs)
        chrome_options.add_argument("--kiosk-printing")
        
        self.driver = webdriver.Chrome(options=chrome_options)
        self.logger.info(f"크롬 드라이버 시작 - 프로필: {profile_dir}")
    
    def convert_to_pdf(self, url, folder, filename):
        """URL을 PDF로 변환"""
        try:
            self.driver.get(url)
            time.sleep(3)  # 페이지 로딩 대기
            
            # PDF로 인쇄
            pdf_path = os.path.join(folder, f"{filename}.pdf")
            result = self.driver.execute_cdp_cmd("Page.printToPDF", {
                "printBackground": True,
                "landscape": False
            })
            
            with open(pdf_path, 'wb') as f:
                f.write(base64.b64decode(result['data']))
            
            return True
        except Exception as e:
            self.logger.error(f"PDF 변환 실패: {url} - {str(e)}")
            return False
    
    def finish_conversion(self):
        """변환 완료 처리"""
        self.is_running = False
        
        # 버튼 상태 복원
        self.start_btn.config(state=NORMAL)
        self.pause_btn.config(state=DISABLED, text="일시정지")
        self.stop_btn.config(state=DISABLED)
        
        self.progress_label.config(text="완료!")
        
        # 결과 요약
        elapsed = int(time.time() - self.start_time)
        minutes = elapsed // 60
        seconds = elapsed % 60
        
        result_msg = f"""
변환 완료!

총 처리: {self.total_processed}개
성공: {self.total_success}개
실패: {self.total_failed}개
소요 시간: {minutes}분 {seconds}초

저장 위치: {self.save_folder}
        """
        
        self.logger.info("=" * 50)
        self.logger.info(result_msg)
        self.logger.info("=" * 50)
        
        # 실패 항목 로그 저장
        if self.failed_items:
            fail_log_path = os.path.join(self.save_folder, "failed_items.txt")
            with open(fail_log_path, 'w', encoding='utf-8') as f:
                f.write("변환 실패 항목\n")
                f.write("=" * 50 + "\n\n")
                for item in self.failed_items:
                    f.write(f"시트: {item['sheet']}\n")
                    f.write(f"행: {item['row']}\n")
                    f.write(f"URL: {item['url']}\n")
                    f.write("-" * 50 + "\n")
            
            result_msg += f"\n\n실패 목록이 저장되었습니다:\n{fail_log_path}"
            self.logger.info(f"실패 목록 저장: {fail_log_path}")
        
        messagebox.showinfo("완료", result_msg)
    
    def cleanup(self):
        """리소스 정리"""
        if self.driver:
            try:
                self.driver.quit()
                self.logger.info("크롬 드라이버 종료")
            except:
                pass
        
        # 크롬 프로세스 강제 종료
        for proc in psutil.process_iter(['name']):
            try:
                if 'chrome' in proc.info['name'].lower():
                    proc.kill()
            except:
                pass


# 간단한 다이얼로그 (tkinter.simpledialog import 대체)
from tkinter import simpledialog

def main():
    root = Tk()
    app = ExcelToPDFApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()
